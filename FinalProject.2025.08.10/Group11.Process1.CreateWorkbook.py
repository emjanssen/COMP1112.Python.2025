# - - - Imports - - - #

import os
import openpyxl
from openpyxl import Workbook

# - - - Function Declarations - - - #

# Create Directory #

def createDirectory():

    adminWorkbookDirectory = r"C:\Users\E.M. Janssen\Documents\GitHub\COMP1112.Python.2025\FinalProject.2025.08.10\Process1.AdminWorkbook"
    if not os.path.exists(adminWorkbookDirectory):
        os.makedirs(adminWorkbookDirectory)
        print("Directory created...")
    else:
        print("Directory already exists...")
    return adminWorkbookDirectory

# Create Admin Workbook #

def createAdminWorkbook(adminWorkbookDirectory):
    from openpyxl import Workbook

    adminWorkbook = Workbook()
    print("Workbook created...")
    return adminWorkbook

def getActiveSheet(adminWorkbook):
     sheetOne = adminWorkbook.active
     print("Got sheet one as active sheet...")
     return sheetOne

def renameSheet(sheetOne):
    sheetOne.title = "Program Team"
    print("Renamed sheet one to 'Program Team'...")
    return sheetOne

def assignColumnHeadings(sheetOne):
    sheetOne['A1'] = "First Name"
    sheetOne['B1'] = "Last Name"
    sheetOne['C1'] = "Company"
    sheetOne['D1'] = "Department"
    sheetOne['E1'] = "Supervisor"
    sheetOne['F1'] = "Start Date"
    sheetOne['G1'] = "End Date"
    sheetOne['H1'] = "Email"
    sheetOne['I1'] = "Landline"
    sheetOne['J1'] = "Mobile"
    sheetOne['K1'] = "Food Allergies"
    print("Input column headings...")

def addSecondSheet(adminWorkbook):
    adminWorkbook.create_sheet("Allergies", 2)
    print("Added second sheet titled 'Allergies'...")
    return adminWorkbook

def saveWorkbook(adminWorkbook, adminWorkbookDirectory):
    filePath = os.path.join(adminWorkbookDirectory, "AdminWorkbook.xlsx")
    adminWorkbook.save(filePath)
    print(f"Workbook saved at {filePath}...")

def returnSheetNames(adminWorkbook):
    adminWorkbookSheetNames = adminWorkbook.sheetnames
    print(f"Our sheet names are: {adminWorkbookSheetNames}")

# - - - Function Calls - - - #

adminWorkbookDirectory = createDirectory()
adminWorkbook = createAdminWorkbook(adminWorkbookDirectory)
sheetOne = getActiveSheet(adminWorkbook)
sheetOne = renameSheet(sheetOne)
assignColumnHeadings(sheetOne)
adminWorkbook = addSecondSheet(adminWorkbook)
saveWorkbook(adminWorkbook, adminWorkbookDirectory)
returnSheetNames(adminWorkbook)