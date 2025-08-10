# - - - Imports - - - #

import os
import openpyxl
from openpyxl.worksheet.copier import WorksheetCopy

# - - - Function Declarations - - - #

def openAdminWorkbook():

    adminWorkbookLocation = r"C:\Users\E.M. Janssen\Documents\GitHub\COMP1112.Python.2025\FinalProject.2025.08.10\Process1.AdminWorkbook\adminWorkbook.xlsx"

    adminWorkbook = openpyxl.load_workbook(adminWorkbookLocation)
    print("We have opened our admin workbook...")
    return adminWorkbook, adminWorkbookLocation

def deleteAllergySheet(adminWorkbook):

    if "Allergies" in adminWorkbook.sheetnames:
        adminWorkbook.remove(adminWorkbook["Allergies"])
        print("Removed old sheet titled 'Allergies'...")
    else:
        print("No sheet named 'Allergies' found.")
    return adminWorkbook

def getActiveSheet(adminWorkbook):

    sheetProgramTeam = adminWorkbook.active
    print(f"Got sheet one as active sheet: {sheetProgramTeam.title}")
    return sheetProgramTeam

def copySheetOne(adminWorkbook, sheetProgramTeam):

    deleteAllergySheet(adminWorkbook)
    allergiesWorksheet = adminWorkbook.create_sheet(title="Allergies")
    worksheetCopyAction = WorksheetCopy(source_worksheet=sheetProgramTeam, target_worksheet=allergiesWorksheet)

    worksheetCopyAction.copy_worksheet()
    print("Copied first sheet into 'Allergies'.")
    return allergiesWorksheet

def setAllergiesSheetAsActive(adminWorkbook, allergiesWorksheet):

    adminWorkbook.active = adminWorkbook.index(allergiesWorksheet)
    print(f"Setting '{allergiesWorksheet.title}' as active sheet...")
    return allergiesWorksheet

def deleteExtraneousColumns(allergiesWorksheet):
    
    columnsToDelete = list(range(3, 11))

    for column in sorted(columnsToDelete, reverse=True):
        allergiesWorksheet.delete_cols(column)

    print("Deleted extraneous columns C to J on Allergies sheet.")
    return allergiesWorksheet

def deleteRowsWithNoAllergyValues(allergiesWorksheet):

    for row in range(allergiesWorksheet.max_row, 1, -1):
        cellValue = allergiesWorksheet.cell(row=row, column=3).value
        
        if cellValue is None or cellValue == "":
            allergiesWorksheet.delete_rows(row)

    print("Deleted rows with no value in the 'Food Allergies' column.")
    return allergiesWorksheet

def saveWorkbook(adminWorkbook):

    adminWorkbookDirectory = r"C:\Users\E.M. Janssen\Documents\GitHub\COMP1112.Python.2025\FinalProject.2025.08.10\Process1.AdminWorkbook"

    filePath = os.path.join(adminWorkbookDirectory, "AdminWorkbook.xlsx")
    adminWorkbook.save(filePath)
    print(f"Workbook saved at {filePath}...")

# - - - Function Calls - - - #

adminWorkbook, adminWorkbookLocation = openAdminWorkbook()
sheetProgramTeam = getActiveSheet(adminWorkbook)
allergiesWorksheet = copySheetOne(adminWorkbook, sheetProgramTeam)
allergiesWorksheet = setAllergiesSheetAsActive(adminWorkbook, allergiesWorksheet)
allergiesWorksheet = deleteExtraneousColumns(allergiesWorksheet)
allergiesWorksheet = deleteRowsWithNoAllergyValues(allergiesWorksheet)
saveWorkbook(adminWorkbook)