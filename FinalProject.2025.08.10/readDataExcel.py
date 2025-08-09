# - - - Imports - - - #

import os
import openpyxl
from openpyxl import Workbook

# - - - Function Declarations - - - #

def readFromUserWorkbook():
    userFilesDirectory = r"C:\Users\E.M. Janssen\Documents\GitHub\COMP1112.Python.2025\FinalProject.2025.08.10\ProgramTeam"
    userFile = "TeamMemberFour.xlsx"
    pathToUserFile = os.path.join(userFilesDirectory, userFile)

    workbookFromUser = openpyxl.load_workbook(pathToUserFile)

    columnTwoText = []

    cellRange = workbookFromUser["Program Info"]["B3":"B13"]
        
    for row in cellRange:
        for cell in row:
            columnTwoText.append(cell.value)
    
    return columnTwoText

def printExcelFileText(columnTwoText):
    print(f"These are our values from the user workbook: {columnTwoText}")

def openAdminWorkbook():

    adminWorkbookLocation = r"C:\Users\E.M. Janssen\Documents\GitHub\COMP1112.Python.2025\FinalProject.2025.08.10\admin\adminWorkbook.xlsx"

    adminWorkbook = openpyxl.load_workbook(adminWorkbookLocation)
    print("We have opened our admin workbook...")
    return adminWorkbook, adminWorkbookLocation

def inputWordDateIntoAdminWorkbook(columnTwoText, adminWorkbook, adminWorkbookLocation):
    
    programTeamSheet = adminWorkbook["Program Team"]

    print(f"Appending row: {columnTwoText}")
    programTeamSheet.append(columnTwoText)
    print("Row appended successfully...")
    
    adminWorkbook.save(adminWorkbookLocation)

# - - - Function Calls - - - #

columnTwoText = readFromUserWorkbook()
printExcelFileText(columnTwoText)
adminWorkbook, adminWorkbookLocation = openAdminWorkbook()
inputWordDateIntoAdminWorkbook(columnTwoText, adminWorkbook, adminWorkbookLocation)