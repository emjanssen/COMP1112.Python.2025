# - - - Imports - - - #

import os
import docx
import openpyxl
from openpyxl import Workbook

# - - - Function Declarations - - - #

def readWordDoc():
    
    # fileFromUser is the variable the admin will change to determine which word doc they wish to read from
    fileFromUser = "TeamMemberTwo.docx"
    folderForUserFiles = r"C:\Users\E.M. Janssen\Documents\GitHub\COMP1112.Python.2025\FinalProject.2025.08.10\ExampleFiles"
    projectRootDirectory = r"C:\Users\E.M. Janssen\Documents\GitHub\COMP1112.Python.2025\FinalProject.2025.08.10"

    fileName = os.path.join(projectRootDirectory, folderForUserFiles, fileFromUser)
    
    document = docx.Document(fileName)
    columnTwoText = []

    for table in document.tables:
        for row in table.rows:
            if len(row.cells) >= 2:
                columnTwoText.append(row.cells[1].text.strip())

    return columnTwoText

def printWordDocText(columnTwoText):
    print(f"This is the content from our word doc: {columnTwoText}")

def openAdminWorkbook():

    adminWorkbookLocation = r"C:\Users\E.M. Janssen\Documents\GitHub\COMP1112.Python.2025\FinalProject.2025.08.10\admin\\adminWorkbook.xlsx"

    adminWorkbook = openpyxl.load_workbook(adminWorkbookLocation)
    print("We have opened our admin workbook...")
    return adminWorkbook, adminWorkbookLocation

def inputWordDateIntoAdminWorkbook(columnTwoText, adminWorkbook, adminWorkbookLocation):
    
    programTeamSheet = adminWorkbook["Program Team"]

    print(f"Appending row: {columnTwoText}")
    programTeamSheet.append(columnTwoText)
    print("Row appended successfully...")
    
    adminWorkbook.save(adminWorkbookLocation)

# - - - Function Calls - - - #

columnTwoText = readWordDoc()
printWordDocText(columnTwoText)
adminWorkbook, adminWorkbookLocation = openAdminWorkbook()
inputWordDateIntoAdminWorkbook(columnTwoText, adminWorkbook, adminWorkbookLocation)